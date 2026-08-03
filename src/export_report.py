#!/usr/bin/env python3
"""
export_report.py - 极速将视频压缩台账导出为专业可视化 Excel (_report.xlsx) 与 HTML 仪表盘 (_report.html)
格式化 MB/Mbps 单位并保留 1 位小数，即刻生成，支持实时动态交互。
"""

import os
import sys
import csv
import json
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def build_excel_report(csv_path, xlsx_path):
    """极速从 CSV 生成专业样式的 Excel 电子表格"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "视频择优压缩台账"
    ws.views.sheetView[0].showGridLines = True

    # 样式定义
    font_family = "PingFang SC"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    
    row_font = Font(name=font_family, size=10)
    succ_font = Font(name=font_family, size=10, color="006100")
    succ_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    skip_font = Font(name=font_family, size=10, color="9C6500")
    skip_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    headers = [
        "相对文件路径", "文件名",
        "压缩前大小(MB)", "原编码", "像素/色彩格式",
        "压缩后大小(MB)", "空间节省率(%)",
        "VMAF画质评分", "测算CRF/Q", "处置决定"
    ]

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
    ws.row_dimensions[1].height = 26

    if not csv_path.exists():
        print(f"Warning: {csv_path} 不存在")
        return

    rows_data = []
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        header_raw = next(reader, None)
        for r in reader:
            if not r or len(r) < 7:
                continue
            rows_data.append(r)

    row_num = 2
    for r in rows_data:
        rel_path = r[0]
        filename = Path(rel_path).name
        
        try:
            orig_mb = round(float(r[1]), 1)
        except ValueError:
            orig_mb = 0.0

        try:
            new_mb = round(float(r[2]), 1)
        except ValueError:
            new_mb = 0.0

        try:
            saving = round(float(r[3]), 1)
        except ValueError:
            saving = 0.0

        vmaf_str = r[4]
        crf_str = r[5]
        decision = r[6]
        pix_fmt = r[7] if len(r) > 7 else ''

        # 写入单元格
        row_vals = [
            rel_path, filename,
            orig_mb, "H.264/MOV" if decision == '保留压缩版' else "HEVC", pix_fmt,
            new_mb if decision == '保留压缩版' else 0.0,
            saving if decision == '保留压缩版' else 0.0,
            vmaf_str, crf_str, decision
        ]
        ws.append(row_vals)

        # 格式与填充
        ws.cell(row=row_num, column=1).alignment = align_left
        ws.cell(row=row_num, column=2).alignment = align_left

        # 数字格式 MB
        ws.cell(row=row_num, column=3).number_format = '0.0 "MB"'
        ws.cell(row=row_num, column=6).number_format = '0.0 "MB"'
        ws.cell(row=row_num, column=7).number_format = '0.0 "%"'

        for c in range(1, len(row_vals) + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.font = row_font
            cell.border = thin_border
            if decision == '保留压缩版':
                cell.fill = succ_fill
                cell.font = succ_font
            elif '跳过' in decision:
                cell.fill = skip_fill
                cell.font = skip_font

        ws.row_dimensions[row_num].height = 20
        row_num += 1

    # 自动列宽
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 32

    wb.save(xlsx_path)
    print(f"✅ 极速 Excel 台账已生成: {xlsx_path}")

def build_html_report(csv_path, html_path):
    """生成具备搜索、排序、卡片统计的可视化 HTML 仪表盘"""
    rows = []
    total_files = 0
    succ_files = 0
    saved_mb_total = 0.0

    if csv_path.exists():
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            header = next(reader, None)
            for r in reader:
                if not r or len(r) < 7:
                    continue
                total_files += 1
                rel_path = r[0]
                filename = Path(rel_path).name
                orig_mb = float(r[1]) if r[1] != 'NA' else 0.0
                new_mb = float(r[2]) if r[2] != 'NA' else 0.0
                saving = float(r[3]) if r[3] != 'NA' else 0.0
                vmaf = r[4]
                crf = r[5]
                decision = r[6]
                pix = r[7] if len(r) > 7 else ''

                if decision == '保留压缩版':
                    succ_files += 1
                    saved_mb_total += (orig_mb - new_mb)

                rows.append({
                    'path': rel_path,
                    'name': filename,
                    'orig_mb': orig_mb,
                    'new_mb': new_mb,
                    'saving': saving,
                    'vmaf': vmaf,
                    'crf': crf,
                    'decision': decision,
                    'pix': pix
                })

    html_content = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <title>vcompress 视频择优压缩可视化台账仪表盘</title>
    <style>
        body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif; background: #f4f6f9; color: #333; margin: 0; padding: 20px; }}
        .header {{ display: flex; justify-content: space-between; align-items: center; background: #fff; padding: 20px 30px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.05); margin-bottom: 20px; }}
        .header h1 {{ margin: 0; font-size: 24px; color: #1f4e79; }}
        .kpi-container {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 15px; margin-bottom: 20px; }}
        .kpi-card {{ background: #fff; padding: 20px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.05); text-align: center; }}
        .kpi-card .num {{ font-size: 28px; font-weight: bold; color: #1f4e79; margin-top: 5px; }}
        .kpi-card .num.green {{ color: #28a745; }}
        .controls {{ display: flex; gap: 15px; margin-bottom: 20px; background: #fff; padding: 15px 20px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.05); }}
        .controls input, .controls select {{ padding: 10px; border: 1px solid #ccc; border-radius: 6px; font-size: 14px; outline: none; }}
        .controls input {{ flex: 1; }}
        table {{ width: 100%; border-collapse: collapse; background: #fff; border-radius: 10px; overflow: hidden; box-shadow: 0 2px 10px rgba(0,0,0,0.05); }}
        th, td {{ padding: 12px 15px; text-align: left; border-bottom: 1px solid #eee; font-size: 14px; }}
        th {{ background: #1f4e79; color: #fff; font-weight: 600; cursor: pointer; }}
        th:hover {{ background: #163857; }}
        tr:hover {{ background: #f8f9fa; }}
        .tag {{ padding: 4px 8px; border-radius: 4px; font-size: 12px; font-weight: bold; display: inline-block; }}
        .tag-succ {{ background: #d4edda; color: #155724; }}
        .tag-skip {{ background: #fff3cd; color: #856404; }}
    </style>
</head>
<body>
    <div class="header">
        <h1>📹 vcompress 视频择优压缩可视化台账仪表盘</h1>
        <span style="color: #666; font-size: 14px;">生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</span>
    </div>

    <div class="kpi-container">
        <div class="kpi-card">
            <div>扫描总视频数</div>
            <div class="num">{total_files}</div>
        </div>
        <div class="kpi-card">
            <div>高性价比压缩成功</div>
            <div class="num green">{succ_files}</div>
        </div>
        <div class="kpi-card">
            <div>累计释放空间</div>
            <div class="num green">{saved_mb_total/1024:.2f} GB</div>
        </div>
        <div class="kpi-card">
            <div>平均节省空间比例</div>
            <div class="num">{round(saved_mb_total / max(sum(r['orig_mb'] for r in rows if r['decision'] == '保留压缩版'), 1.0) * 100, 1):.1f}%</div>
        </div>
    </div>

    <div class="controls">
        <input type="text" id="searchInput" onkeyup="filterTable()" placeholder="🔍 搜索文件名或路径...">
        <select id="statusFilter" onchange="filterTable()">
            <option value="ALL">全部结果决定</option>
            <option value="保留压缩版">保留压缩版 (已压缩)</option>
            <option value="跳过">跳过 (未压缩)</option>
        </select>
    </div>

    <table id="reportTable">
        <thead>
            <tr>
                <th onclick="sortTable(0)">文件名</th>
                <th onclick="sortTable(1)">原大小 (MB)</th>
                <th onclick="sortTable(2)">压缩后 (MB)</th>
                <th onclick="sortTable(3)">节省比例</th>
                <th onclick="sortTable(4)">VMAF评分</th>
                <th onclick="sortTable(5)">最佳CRF</th>
                <th onclick="sortTable(6)">处置决定</th>
                <th onclick="sortTable(7)">相对路径</th>
            </tr>
        </thead>
        <tbody>
"""

    for r in rows:
        tag_cls = "tag-succ" if r['decision'] == '保留压缩版' else "tag-skip"
        new_mb_disp = f"{r['new_mb']:.1f} MB" if r['decision'] == '保留压缩版' else "-"
        saving_disp = f"{r['saving']:.1f}%" if r['decision'] == '保留压缩版' else "-"

        html_content += f"""
            <tr>
                <td><strong>{r['name']}</strong></td>
                <td>{r['orig_mb']:.1f} MB</td>
                <td>{new_mb_disp}</td>
                <td>{saving_disp}</td>
                <td>{r['vmaf']}</td>
                <td>{r['crf']}</td>
                <td><span class="tag {tag_cls}">{r['decision']}</span></td>
                <td style="color:#666; font-size:12px;">{r['path']}</td>
            </tr>
"""

    html_content += """
        </tbody>
    </table>

    <script>
        function filterTable() {
            var input = document.getElementById("searchInput").value.toUpperCase();
            var status = document.getElementById("statusFilter").value;
            var tr = document.getElementById("reportTable").getElementsByTagName("tr");

            for (var i = 1; i < tr.length; i++) {
                var tdName = tr[i].getElementsByTagName("td")[0];
                var tdPath = tr[i].getElementsByTagName("td")[7];
                var tdStatus = tr[i].getElementsByTagName("td")[6];
                
                if (tdName && tdStatus) {
                    var txtValue = (tdName.textContent || tdName.innerText) + " " + (tdPath.textContent || tdPath.innerText);
                    var statusValue = tdStatus.textContent || tdStatus.innerText;

                    var matchesSearch = txtValue.toUpperCase().indexOf(input) > -1;
                    var matchesStatus = (status === "ALL") || 
                                        (status === "保留压缩版" && statusValue.indexOf("保留压缩版") > -1) ||
                                        (status === "跳过" && statusValue.indexOf("保留压缩版") === -1);

                    if (matchesSearch && matchesStatus) {
                        tr[i].style.display = "";
                    } else {
                        tr[i].style.display = "none";
                    }
                }
            }
        }

        function sortTable(n) {
            var table = document.getElementById("reportTable");
            var rows, switching, i, x, y, shouldSwitch, dir, switchcount = 0;
            switching = true;
            dir = "asc";
            while (switching) {
                switching = false;
                rows = table.rows;
                for (i = 1; i < (rows.length - 1); i++) {
                    shouldSwitch = false;
                    x = rows[i].getElementsByTagName("TD")[n];
                    y = rows[i + 1].getElementsByTagName("TD")[n];

                    var xVal = x.textContent || x.innerText;
                    var yVal = y.textContent || y.innerText;

                    var xNum = parseFloat(xVal.replace(/[^0-9.-]+/g,""));
                    var yNum = parseFloat(yVal.replace(/[^0-9.-]+/g,""));

                    if (!isNaN(xNum) && !isNaN(yNum)) {
                        if (dir == "asc") {
                            if (xNum > yNum) { shouldSwitch = true; break; }
                        } else if (dir == "desc") {
                            if (xNum < yNum) { shouldSwitch = true; break; }
                        }
                    } else {
                        if (dir == "asc") {
                            if (xVal.toLowerCase() > yVal.toLowerCase()) { shouldSwitch = true; break; }
                        } else if (dir == "desc") {
                            if (xVal.toLowerCase() < yVal.toLowerCase()) { shouldSwitch = true; break; }
                        }
                    }
                }
                if (shouldSwitch) {
                    rows[i].parentNode.insertBefore(rows[i + 1], rows[i]);
                    switching = true;
                    switchcount ++;
                } else {
                    if (switchcount == 0 && dir == "asc") {
                        dir = "desc";
                        switching = true;
                    }
                }
            }
        }
    </script>
</body>
</html>
"""
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html_content)
    print(f"✅ 可视化 HTML 仪表盘台账已保存: {html_path}")

def main():
    if len(sys.argv) < 2:
        print("Usage: python3 export_report.py <csv_path_or_output_dir>")
        sys.exit(1)

    target_path = Path(sys.argv[1])
    if target_path.is_dir():
        csv_path = target_path / "_report.csv"
        out_dir = target_path
    else:
        csv_path = target_path
        out_dir = target_path.parent

    xlsx_path = out_dir / "_report.xlsx"
    html_path = out_dir / "_report.html"

    print(f"读取 CSV 台账: {csv_path}...")
    build_excel_report(csv_path, xlsx_path)
    build_html_report(csv_path, html_path)

if __name__ == "__main__":
    main()
